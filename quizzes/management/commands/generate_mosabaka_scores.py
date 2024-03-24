from django.core.management.base import BaseCommand
from openpyxl import Workbook
from quizzes.models import MosabakaScore
from quizzes.serializers import MosabakaScoreSerializer

class Command(BaseCommand):
    help = 'Generate an Excel file with all MosabakaScore data'

    def handle(self, *args, **options):
        queryset = MosabakaScore.objects.all().order_by('-score')
        serializer = MosabakaScoreSerializer(queryset, many=True)
        mosabaka_scores = serializer.data

        # Create a new Excel workbook
        workbook = Workbook()
        worksheet = workbook.active

        # Write the column headers
        worksheet['A1'] = 'الإسم الكامل'
        worksheet['B1'] = 'المسابقة'
        worksheet['C1'] = 'النتيجة'

        # Write the data to the worksheet
        row = 2
        for score in mosabaka_scores:
            worksheet.cell(row=row, column=1, value=score['student_name'])
            worksheet.cell(row=row, column=2, value=score['mosabaka_name'])
            worksheet.cell(row=row, column=3, value=score['score'])
            row += 1

        # Save the workbook to a file
        workbook.save('mosabaka_scores.xlsx')

        self.stdout.write(self.style.SUCCESS('Excel file generated successfully'))